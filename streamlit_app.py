import streamlit as st
import gspread
import pandas as pd
from google.oauth2.service_account import Credentials
from datetime import datetime
from dateutil.relativedelta import relativedelta
import uuid
import re
import numpy as np
import plotly.express as px
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# =============================================================================
# 0. 기본 설정 및 상수 정의
# =============================================================================
st.set_page_config(page_title="통합 정산 관리 시스템", page_icon="🏢", layout="wide")

SHEET_NAMES = {
    "SETTINGS": "시스템_설정", "LOCATIONS": "사업장_마스터", "ACCOUNTS": "계정과목_마스터",
    "RULES": "자동분류_규칙", "TRANSACTIONS": "통합거래_원장", "INVENTORY": "월별재고_자산",
    "FORMATS": "파일_포맷_마스터"
}

# 파싱 상수 정의
OKPOS_DATA_START_ROW, OKPOS_COL_DATE, OKPOS_COL_DINE_IN, OKPOS_COL_TAKEOUT, OKPOS_COL_DELIVERY = 7, 0, 34, 36, 38
WOORI_DATA_START_ROW, WOORI_COL_CHECK, WOORI_COL_DATETIME, WOORI_COL_DESC, WOORI_COL_AMOUNT = 4, 0, 1, 3, 4

# =============================================================================
# ★★★ 전용 파서 및 헬퍼 함수들 ★★★
# =============================================================================
def parse_okpos(df_raw):
    out = []
    try:
        end_row_series = df_raw[df_raw.iloc[:, OKPOS_COL_DATE].astype(str).str.contains("합계", na=False)].index
        end_row = end_row_series[0] if not end_row_series.empty else df_raw.shape[0]
    except Exception: end_row = df_raw.shape[0]
    df_data = df_raw.iloc[OKPOS_DATA_START_ROW:end_row]
    for _, row in df_data.iterrows():
        try:
            date_cell = row.iloc[OKPOS_COL_DATE]
            if pd.isna(date_cell): continue
            cleaned_date_str = str(date_cell).replace("소계:", "").strip()
            date = pd.to_datetime(cleaned_date_str).strftime('%Y-%m-%d')
            홀매출 = pd.to_numeric(row.iloc[OKPOS_COL_DINE_IN], errors='coerce')
            포장매출 = pd.to_numeric(row.iloc[OKPOS_COL_TAKEOUT], errors='coerce')
            배달매출 = pd.to_numeric(row.iloc[OKPOS_COL_DELIVERY], errors='coerce')
            if pd.notna(홀매출) and 홀매출 != 0: out.append({'거래일자': date, '거래내용': 'OKPOS 홀매출', '금액': 홀매출})
            if pd.notna(포장매출) and 포장매출 != 0: out.append({'거래일자': date, '거래내용': 'OKPOS 포장매출', '금액': 포장매출})
            if pd.notna(배달매출) and 배달매출 != 0: out.append({'거래일자': date, '거래내용': 'OKPOS 배달매출', '금액': 배달매출})
        except Exception: continue
    return pd.DataFrame(out)

def parse_woori_bank(df_raw):
    out, error_rows = [], []
    df_data = df_raw.iloc[WOORI_DATA_START_ROW:].copy()
    for index, row in df_data.iterrows():
        excel_row_num = index + 1
        try:
            check_val = row.iloc[WOORI_COL_CHECK]
            if pd.isna(pd.to_numeric(check_val, errors='coerce')): break
            datetime_str = str(row.iloc[WOORI_COL_DATETIME]).split(' ')[0]
            date = pd.to_datetime(datetime_str).strftime('%Y-%m-%d')
            description = str(row.iloc[WOORI_COL_DESC])
            amount_str = str(row.iloc[WOORI_COL_AMOUNT]).replace(',', '')
            amount = pd.to_numeric(amount_str, errors='coerce')
            if pd.notna(amount) and amount > 0 and description.strip() != '':
                out.append({'거래일자': date, '거래내용': description, '금액': amount})
            else: error_rows.append(excel_row_num)
        except Exception: error_rows.append(excel_row_num); continue
    if error_rows: st.warning(f"⚠️ **{len(error_rows)}개 행 변환 누락:** 원본 파일의 다음 행들을 확인해주세요: {', '.join(map(str, error_rows[:10]))}{'...' if len(error_rows) > 10 else ''}")
    return pd.DataFrame(out)

# =============================================================================
# 1. 구글 시트 연결
# =============================================================================
def get_spreadsheet_key():
    try: return st.secrets["gcp_service_account"]["SPREADSHEET_KEY"]
    except KeyError:
        try: return st.secrets["SPREADSHEET_KEY"]
        except KeyError: st.error("Streamlit Secrets에 'SPREADSHEET_KEY'를 찾을 수 없습니다."); st.stop()

@st.cache_resource
def get_gspread_client():
    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive.file"]
    creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=scopes)
    return gspread.authorize(creds)

@st.cache_data(ttl=60)
def load_data(sheet_name):
    try:
        spreadsheet_key = get_spreadsheet_key()
        spreadsheet = get_gspread_client().open_by_key(spreadsheet_key)
        worksheet = spreadsheet.worksheet(sheet_name)
        df = pd.DataFrame(worksheet.get_all_records(head=1))
        for col in df.columns: df[col] = df[col].astype(str).str.strip()
        numeric_cols = ['금액', '기말재고액']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col].str.replace(',', ''), errors='coerce').fillna(0)
        return df
    except gspread.exceptions.WorksheetNotFound: st.error(f"'{sheet_name}' 시트를 찾을 수 없습니다."); return pd.DataFrame()
    except Exception as e: st.error(f"'{sheet_name}' 시트 로딩 중 오류: {e}"); return pd.DataFrame()

# --- 핵심 수정: 시트가 비워지는 오류를 막기 위한 안정성 강화 ---
def update_sheet(sheet_name, df):
    try:
        spreadsheet_key = get_spreadsheet_key()
        spreadsheet = get_gspread_client().open_by_key(spreadsheet_key)
        worksheet = spreadsheet.worksheet(sheet_name)
        
        # 데이터프레임이 비어있더라도 헤더는 항상 존재하도록 보장
        header = df.columns.values.tolist()
        
        # 저장할 데이터가 있을 때만 시트를 지우고 새로 씀
        if not df.empty:
            worksheet.clear()
            if '거래일자' in df.columns:
                df['거래일자'] = pd.to_datetime(df['거래일자']).dt.strftime('%Y-%m-%d')
            df_str = df.astype(str).replace('nan', '').replace('NaT', '')
            worksheet.update([header] + df_str.values.tolist(), value_input_option='USER_ENTERED')
        else:
            # 데이터가 없으면 헤더만 다시 씀 (내용이 모두 삭제된 경우)
            worksheet.clear()
            worksheet.update([header], value_input_option='USER_ENTERED')
            
        st.cache_data.clear()
        return True
    except Exception as e:
        st.error(f"'{sheet_name}' 시트 업데이트 중 오류: {e}")
        return False

# =============================================================================
# 2. 로그인, 3. 핵심 로직
# =============================================================================
def login_screen():
    st.title("🏢 통합 정산 관리 시스템")
    settings_df = load_data(SHEET_NAMES["SETTINGS"])
    if settings_df.empty: st.error("`시스템_설정` 시트가 비어있습니다."); st.stop()
    admin_id_row = settings_df[settings_df['Key'] == 'ADMIN_ID']
    admin_pw_row = settings_df[settings_df['Key'] == 'ADMIN_PW']
    if admin_id_row.empty or admin_pw_row.empty: st.error("`시스템_설정` 시트에 ADMIN_ID/PW Key가 없습니다."); st.stop()
    admin_id, admin_pw = admin_id_row['Value'].iloc[0], admin_pw_row['Value'].iloc[0]
    with st.form("login_form"):
        username, password = st.text_input("아이디"), st.text_input("비밀번호", type="password")
        if st.form_submit_button("로그인", use_container_width=True):
            if username == admin_id and password == admin_pw: st.session_state['logged_in'] = True; st.rerun()
            else: st.error("아이디 또는 비밀번호가 올바라지 않습니다.")

def auto_categorize(df, rules_df):
    if rules_df.empty: return df
    categorized_df = df.copy()
    for index, row in categorized_df.iterrows():
        if pd.notna(row.get('계정ID')) and row.get('계정ID') != '': continue
        description = str(row['거래내용'])
        for _, rule in rules_df.iterrows():
            keyword = str(rule['키워드'])
            if keyword and keyword in description:
                categorized_df.loc[index, '계정ID'] = rule['계정ID']
                categorized_df.loc[index, '처리상태'] = '자동분류'; break
    return categorized_df

def calc_change(current, prev):
    if prev > 0:
        return ((current - prev) / prev) * 100
    return np.inf if current > 0 else 0

def calculate_pnl_new(transactions_df, accounts_df, selected_month, selected_location):
    if transactions_df.empty or '거래일자' not in transactions_df.columns:
        return {}, pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    def get_monthly_data(month_str):
        month_trans = transactions_df[transactions_df['거래일자'].dt.strftime('%Y-%m') == month_str].copy()
        if month_trans.empty:
            return {'총매출': 0, '총비용': 0, '영업이익': 0}, pd.DataFrame(columns=['소분류', '금액']), pd.DataFrame(columns=['대분류', '소분류', '금액']), pd.DataFrame()

        pnl_data = pd.merge(month_trans, accounts_df, on='계정ID', how='left')
        pnl_data['대분류'] = pnl_data['대분류'].fillna('기타')
        
        sales_df = pnl_data[pnl_data['대분류'].str.contains('매출', na=False)]
        total_sales = sales_df['금액'].sum()
        
        expenses_df = pnl_data[~pnl_data['대분류'].str.contains('매출', na=False)]
        total_expenses = expenses_df['금액'].sum()
        
        operating_profit = total_sales - total_expenses
        
        metrics = {"총매출": total_sales, "총비용": total_expenses, "영업이익": operating_profit}
        sales_breakdown = sales_df.groupby('소분류')['금액'].sum().reset_index()
        expense_breakdown = expenses_df.groupby(['대분류', '소분류'])['금액'].sum().reset_index()
        
        return metrics, sales_breakdown, expense_breakdown, pnl_data

    transactions_df['거래일자'] = pd.to_datetime(transactions_df['거래일자'], errors='coerce')

    if selected_location != "전체":
        transactions_df = transactions_df[transactions_df['사업장명'] == selected_location]

    prev_month_str = (datetime.strptime(selected_month + '-01', '%Y-%m-%d') - relativedelta(months=1)).strftime('%Y-%m')
    
    current_metrics, current_sales, current_expenses, current_details = get_monthly_data(selected_month)
    prev_metrics, _, prev_expenses, _ = get_monthly_data(prev_month_str)
    
    current_metrics['총매출_증감'] = calc_change(current_metrics['총매출'], prev_metrics['총매출'])
    current_metrics['총비용_증감'] = calc_change(current_metrics['총비용'], prev_metrics['총비용'])
    current_metrics['영업이익_증감'] = calc_change(current_metrics['영업이익'], prev_metrics['영업이익'])
    current_metrics['영업이익률'] = (current_metrics['영업이익'] / current_metrics['총매출']) * 100 if current_metrics['총매출'] > 0 else 0
    
    if not current_expenses.empty:
        expense_merged = pd.merge(current_expenses, prev_expenses, on=['대분류', '소분류'], how='outer', suffixes=('_현재', '_과거')).fillna(0)
        expense_merged['증감률'] = expense_merged.apply(lambda row: calc_change(row['금액_현재'], row['금액_과거']), axis=1)
    else:
        expense_merged = pd.DataFrame(columns=['대분류', '소분류', '금액_현재', '금액_과거', '증감률'])

    return current_metrics, current_sales, expense_merged, current_details

def create_excel_report(selected_month, selected_location, metrics, sales_breakdown, expense_breakdown, pnl_details_df):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "손익계산서 대시보드"

    title_font = Font(name='맑은 고딕', size=16, bold=True)
    header_font = Font(name='맑은 고딕', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def apply_header_style(worksheet, start_row, start_col, end_col):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=start_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
    
    def auto_fit_columns(worksheet):
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width

    ws.merge_cells('B2:F2')
    title_cell = ws['B2']
    title_cell.value = f"{selected_month} 월별 정산표 ({selected_location})"
    title_cell.font = title_font
    title_cell.alignment = center_align

    summary_data = [
        ["항목", "당월 금액", "전월 대비 증감률(%)"],
        ["총매출", metrics['총매출'], f"{metrics['총매출_증감']:.1f}%"],
        ["총비용", metrics['총비용'], f"{metrics['총비용_증감']:.1f}%"],
        ["영업이익", metrics['영업이익'], f"{metrics['영업이익_증감']:.1f}%"],
        ["영업이익률", f"{metrics['영업이익률']:.1f}%", ""]
    ]
    current_row = 4
    for r_idx, row_data in enumerate(summary_data, current_row):
        ws.append([""] + row_data)
    apply_header_style(ws, 5, 2, 4)

    current_row = ws.max_row + 2
    ws.cell(row=current_row, column=2, value="매출 상세").font = title_font
    sales_df_rows = dataframe_to_rows(sales_breakdown, index=False, header=True)
    for r_idx, row in enumerate(sales_df_rows, current_row + 1):
        ws.append([""] + row)
    apply_header_style(ws, current_row + 1, 2, 2 + sales_breakdown.shape[1] - 1)

    current_row = ws.max_row + 2
    ws.cell(row=current_row, column=2, value="비용 상세").font = title_font
    expense_report_df = expense_breakdown.rename(columns={'금액_현재': '당월 금액', '금액_과거': '전월 금액', '증감률': '증감률(%)'})
    expense_df_rows = dataframe_to_rows(expense_report_df[['대분류', '소분류', '당월 금액', '전월 금액', '증감률(%)']], index=False, header=True)
    for r_idx, row in enumerate(expense_df_rows, current_row + 1):
        ws.append([""] + row)
    apply_header_style(ws, current_row + 1, 2, 2 + expense_report_df.shape[1] - 1)
    
    auto_fit_columns(ws)

    ws2 = wb.create_sheet("세부 거래 내역")
    if not pnl_details_df.empty:
        detail_cols = ['거래일자', '사업장명', '대분류', '소분류', '거래내용', '금액']
        df_details_final = pnl_details_df[detail_cols].sort_values(by="거래일자")
        for r in dataframe_to_rows(df_details_final, index=False, header=True):
            ws2.append(r)
        
        apply_header_style(ws2, 1, 1, len(detail_cols))
        auto_fit_columns(ws2)
    
    wb.save(output)
    return output.getvalue()

def calculate_trend_data(transactions_df, accounts_df, end_month_str, num_months, selected_location):
    if transactions_df.empty or '거래일자' not in transactions_df.columns:
        return pd.DataFrame()

    trend_data = []
    end_month = datetime.strptime(end_month_str + '-01', '%Y-%m-%d')
    
    if selected_location != "전체":
        transactions_df = transactions_df[transactions_df['사업장명'] == selected_location]

    transactions_df['거래일자'] = pd.to_datetime(transactions_df['거래일자'], errors='coerce')
    
    for i in range(num_months - 1, -1, -1):
        month = end_month - relativedelta(months=i)
        month_str = month.strftime('%Y-%m')
        
        month_trans = transactions_df[transactions_df['거래일자'].dt.strftime('%Y-%m') == month_str]
        pnl_data = pd.merge(month_trans, accounts_df, on='계정ID', how='left')
        pnl_data['대분류'] = pnl_data['대분류'].fillna('기타')
        
        total_sales = pnl_data[pnl_data['대분류'].str.contains('매출', na=False)]['금액'].sum()
        total_expenses = pnl_data[~pnl_data['대분류'].str.contains('매출', na=False)]['금액'].sum()
        
        trend_data.append({'월': month_str, '총매출': total_sales, '총비용': total_expenses})
        
    return pd.DataFrame(trend_data)

# =============================================================================
# 4. UI 렌더링 함수
# =============================================================================
def render_pnl_page(data):
    st.header("📅 월별 정산표")

    top_col1, top_col2, top_col3 = st.columns([0.35, 0.35, 0.3])
    location_list = ["전체"] + data["LOCATIONS"]['사업장명'].tolist() if not data["LOCATIONS"].empty else ["전체"]
    selected_location = top_col1.selectbox("사업장 선택", location_list)
    month_options = [(datetime.now() - relativedelta(months=i)).strftime('%Y-%m') for i in range(12)]
    selected_month = top_col2.selectbox("조회 년/월 선택", month_options)
    view_option = top_col3.selectbox("보기 옵션", ["월별 상세 보기", "매출/비용 추세"], index=0)

    if not selected_month: st.stop()
    st.markdown("---")
    
    if view_option == "월별 상세 보기":
        metrics, sales_breakdown, expense_breakdown, pnl_details_df = calculate_pnl_new(data["TRANSACTIONS"], data["ACCOUNTS"], selected_month, selected_location)

        if not metrics or (metrics['총매출'] == 0 and metrics['총비용'] == 0):
            st.warning(f"'{selected_location}'의 {selected_month} 데이터가 없습니다."); st.stop()
        
        excel_data = create_excel_report(selected_month, selected_location, metrics, sales_breakdown, expense_breakdown, pnl_details_df)
        st.download_button("📥 엑셀로 다운로드", excel_data, f"{selected_month}_{selected_location}_정산표.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        st.markdown("---")

        summary_col, chart_col = st.columns([0.6, 0.4])
        with summary_col:
            st.subheader("📊 손익 요약")
            m1, m2, m3 = st.columns(3)
            m1.metric("총매출", f"{metrics['총매출']:,.0f} 원", f"{metrics['총매출_증감']:.1f}%")
            m2.metric("총비용", f"{metrics['총비용']:,.0f} 원", f"{metrics['총비용_증감']:.1f}%", delta_color="inverse")
            m3.metric("영업이익", f"{metrics['영업이익']:,.0f} 원", f"{metrics['영업이익_증감']:.1f}%")
            st.markdown("---")

            with st.expander(f"**Ⅰ. 총매출: {metrics['총매출']:,.0f} 원**", expanded=False):
                st.dataframe(sales_breakdown.rename(columns={'소분류': '항목', '금액': '금액(원)'}), use_container_width=True, hide_index=True)

            with st.expander(f"**Ⅱ. 총비용: {metrics['총비용']:,.0f} 원**", expanded=True):
                expense_order = ['인건비', '식자재', '소모품', '광고비', '고정비']
                all_major_cats = expense_breakdown['대분류'].unique()
                sorted_major_cats = [cat for cat in expense_order if cat in all_major_cats] + [cat for cat in all_major_cats if cat not in expense_order and cat != 0]

                for major_cat in sorted_major_cats:
                    major_df = expense_breakdown[expense_breakdown['대분류'] == major_cat]
                    major_total_current = major_df['금액_현재'].sum()
                    major_total_prev = major_df['금액_과거'].sum()
                    major_mom = calc_change(major_total_current, major_total_prev)
                    major_percentage = (major_total_current / metrics['총비용']) * 100 if metrics['총비용'] > 0 else 0
                    
                    delta_str = f"{major_mom:+.1f}%" if np.isfinite(major_mom) else "N/A"
                    expander_title = f"**{major_cat}: {major_total_current:,.0f} 원 ({major_percentage:.1f}%)**"
                    
                    with st.expander(expander_title):
                        st.caption(f"전월 대비: {delta_str}")
                        for _, row in major_df.iterrows():
                            sub_col1, sub_col2, sub_col3 = st.columns([0.6, 0.2, 0.2])
                            sub_col1.markdown(f"- {row['소분류']}: **{row['금액_현재']:,.0f} 원**")
                            delta_text = f"{row['증감률']:+.1f}%" if np.isfinite(row['증감률']) else ""
                            sub_col2.metric("", "", delta=delta_text, delta_color="inverse")
                            if sub_col3.button("거래 보기", key=f"btn_{row['소분류']}", use_container_width=True):
                                detail_df = pnl_details_df[pnl_details_df['소분류'] == row['소분류']]
                                st.dataframe(detail_df[['거래일자', '사업장명', '거래내용', '금액']].sort_values('거래일자'), use_container_width=True, hide_index=True)
            
            st.markdown(f"--- \n ### **Ⅲ. 영업이익: {metrics['영업이익']:,.0f} 원 ({metrics['영업이익률']:.1f}%)**")

        with chart_col:
            st.subheader("📈 시각화 분석")
            if not sales_breakdown.empty:
                st.markdown("**매출 비중**")
                fig_pie_sales = px.pie(sales_breakdown, names='소분류', values='금액', hole=.4, title=f"총 매출: {metrics['총매출']:,.0f} 원")
                fig_pie_sales.update_traces(textinfo='percent+label', textfont_size=14)
                st.plotly_chart(fig_pie_sales, use_container_width=True)
            
            if not expense_breakdown.empty:
                expense_by_major = expense_breakdown.groupby('대분류')['금액_현재'].sum().reset_index()
                st.markdown("**비용 비중**")
                fig_pie_expenses = px.pie(expense_by_major, names='대분류', values='금액_현재', hole=.4, title=f"총 비용: {metrics['총비용']:,.0f} 원")
                fig_pie_expenses.update_traces(textinfo='percent+label', textfont_size=14)
                st.plotly_chart(fig_pie_expenses, use_container_width=True)
    
    elif view_option == "매출/비용 추세":
        st.subheader(f"📈 {selected_month} 기준, 최근 데이터 추세")
        period = st.radio("기간 선택", [3, 6, 12], index=1, horizontal=True)
        trend_df = calculate_trend_data(data["TRANSACTIONS"], data["ACCOUNTS"], selected_month, period, selected_location)
        
        if trend_df.empty:
            st.warning("추세 데이터를 표시할 정보가 부족합니다.")
        else:
            st.line_chart(trend_df.set_index('월'))
            st.dataframe(trend_df, use_container_width=True, hide_index=True)

def render_data_page(data):
    st.header("✍️ 데이터 관리")
    if 'current_step' not in st.session_state:
        st.session_state.current_step = 'upload'
    if st.session_state.current_step == 'upload':
        st.subheader("🏢 데이터 현황")
        if data["TRANSACTIONS"].empty:
            st.info("아직 등록된 거래내역이 없습니다. 아래에서 파일을 업로드해주세요.")
        else:
            trans_df_copy = data["TRANSACTIONS"].copy()
            trans_df_copy['거래일자'] = pd.to_datetime(trans_df_copy['거래일자'], errors='coerce').dt.normalize()
            summary = trans_df_copy.groupby(['사업장명', '데이터소스']).agg(건수=('거래ID', 'count'), 최초거래일=('거래일자', 'min'), 최종거래일=('거래일자', 'max')).reset_index()
            for location in data["LOCATIONS"]['사업장명']:
                st.markdown(f"**{location}**")
                loc_summary = summary[summary['사업장명'] == location]
                if loc_summary.empty:
                    st.write("└ 데이터 없음")
                else:
                    for _, row in loc_summary.iterrows():
                        st.write(f"└ `{row['데이터소스']}`: {row['최초거래일'].strftime('%Y-%m-%d')} ~ {row['최종거래일'].strftime('%Y-%m-%d')} (총 {row['건수']}건)")
        st.markdown("---")
        if data["LOCATIONS"].empty or data["ACCOUNTS"].empty or data["FORMATS"].empty:
            st.error("`설정 관리`에서 `사업장`, `계정과목`, `파일 포맷`을 먼저 등록해야 합니다.")
            st.stop()
        tab1, tab2 = st.tabs(["거래내역 관리 (파일 업로드)", "월별재고 관리"])
        with tab1:
            st.subheader("파일 기반 거래내역 관리")
            format_list = data["FORMATS"]['포맷명'].tolist()
            selected_format_name = st.selectbox("1. 처리할 파일 포맷을 선택하세요.", format_list)
            location_list = data["LOCATIONS"]['사업장명'].tolist()
            upload_location = st.selectbox("2. 데이터를 귀속시킬 사업장을 선택하세요.", location_list)
            uploaded_file = st.file_uploader("3. 해당 포맷의 파일을 업로드하세요.", type=["xlsx", "xls", "csv"])
            if st.button("4. 파일 처리 및 데이터 확인", type="primary", use_container_width=True):
                if not uploaded_file:
                    st.error("파일을 먼저 업로드해주세요.")
                else:
                    with st.spinner("파일을 처리하는 중입니다..."):
                        df_raw = None
                        try:
                            if uploaded_file.name.endswith('.csv'):
                                try:
                                    df_raw = pd.read_csv(uploaded_file, encoding='utf-8', header=None)
                                except UnicodeDecodeError:
                                    uploaded_file.seek(0)
                                    df_raw = pd.read_csv(uploaded_file, encoding='cp949', header=None)
                            else:
                                df_raw = pd.read_excel(uploaded_file, header=None)
                        except Exception as e:
                            st.error(f"파일을 읽는 중 오류가 발생했습니다: {e}")
                            return
                        if df_raw is None:
                            st.error("지원하지 않는 파일 형식입니다.")
                            return
                        df_parsed = pd.DataFrame()
                        if selected_format_name == "OKPOS 매출":
                            df_parsed = parse_okpos(df_raw)
                        elif selected_format_name == "우리은행 지출":
                            df_parsed = parse_woori_bank(df_raw)
                        if df_parsed.empty:
                            st.warning("파일에서 처리할 데이터를 찾지 못했습니다.")
                            return
                        df_final = df_parsed.copy()
                        df_final['사업장명'] = upload_location
                        df_final['구분'] = data["FORMATS"][data["FORMATS"]['포맷명'] == selected_format_name].iloc[0]['데이터구분']
                        df_final['데이터소스'] = selected_format_name
                        df_final['거래ID'] = [str(uuid.uuid4()) for _ in range(len(df_final))]
                        st.session_state.uploaded_file_metadata = {'사업장명': upload_location, '구분': df_final['구분'].iloc[0], '데이터소스': selected_format_name}
                        if selected_format_name == "OKPOS 매출":
                            def get_okpos_account_id(description):
                                accounts_df = data["ACCOUNTS"]
                                account_id = accounts_df.loc[accounts_df['소분류'] == description, '계정ID']
                                return account_id.iloc[0] if not account_id.empty else ''
                            df_final['계정ID'] = df_final['거래내용'].apply(get_okpos_account_id)
                            df_final['처리상태'] = '자동등록'
                            st.session_state.okpos_preview_data = df_final
                            st.session_state.current_step = 'okpos_preview'
                        else:
                            df_final['처리상태'] = '미분류'
                            df_final['계정ID'] = ''
                            st.session_state.df_processed = df_final
                            st.session_state.current_step = 'confirm'
                        st.rerun()
        with tab2:
            st.subheader("월별재고 관리")
            if data["LOCATIONS"].empty:
                st.warning("`설정 관리` 탭에서 `사업장`을 먼저 추가해주세요.")
            else:
                edited_inv = st.data_editor(data["INVENTORY"], num_rows="dynamic", use_container_width=True, hide_index=True, column_config={"사업장명": st.column_config.SelectboxColumn("사업장명", options=data["LOCATIONS"]['사업장명'].tolist(), required=True)})
                if st.button("💾 월별재고 저장", key="save_inventory"):
                    if update_sheet(SHEET_NAMES["INVENTORY"], edited_inv):
                        st.success("저장되었습니다.")
                        st.rerun()
    elif st.session_state.current_step == 'okpos_preview':
        st.subheader("✅ OKPOS 매출 데이터 미리보기 및 저장")
        df_preview = st.session_state.get('okpos_preview_data', pd.DataFrame())
        if df_preview.empty:
            st.warning("미리보기할 데이터가 없습니다. 이전 단계로 돌아가세요.")
        else:
            st.dataframe(df_preview[['거래일자', '거래내용', '금액', '계정ID']], use_container_width=True, hide_index=True)
        col1, col2 = st.columns(2)
        if col1.button("🔙 이전 단계로"):
            del st.session_state.okpos_preview_data
            st.session_state.current_step = 'upload'
            st.rerun()
        if col2.button("💾 최종 저장하기", type="primary"):
            if (df_preview['계정ID'] == '').any():
                st.error("계정과목_마스터에 OKPOS 매출 항목(OKPOS 홀매출 등)이 등록되지 않았거나, 이름이 다릅니다.")
            else:
                with st.spinner("데이터를 저장하는 중입니다..."):
                    combined_trans = pd.concat([data["TRANSACTIONS"], df_preview], ignore_index=True)
                    if update_sheet(SHEET_NAMES["TRANSACTIONS"], combined_trans):
                        st.success(f"OKPOS 매출 데이터 {len(df_preview)}건이 성공적으로 저장되었습니다.")
                        del st.session_state.okpos_preview_data
                        st.session_state.current_step = 'upload'
                        st.rerun()
    elif st.session_state.current_step == 'confirm':
        st.subheader("✅ 1단계: 확인 및 확정")
        df_processed = st.session_state.get('df_processed', pd.DataFrame())
        df_non_duplicates = df_processed.copy()
        df_duplicates = pd.DataFrame()
        if not df_processed.empty and df_processed['구분'].iloc[0] == '비용':
            existing = data["TRANSACTIONS"]
            if not existing.empty:
                existing['duplicate_key'] = existing['사업장명'] + existing['거래내용'] + existing['금액'].astype(str)
                df_processed['duplicate_key'] = df_processed['사업장명'] + df_processed['거래내용'] + df_processed['금액'].astype(str)
                is_duplicate = df_processed['duplicate_key'].isin(existing['duplicate_key'])
                df_duplicates = df_processed[is_duplicate]
                df_non_duplicates = df_processed[~is_duplicate]
        df_processed_no_duplicates = auto_categorize(df_non_duplicates, data["RULES"])
        df_auto = df_processed_no_duplicates[df_processed_no_duplicates['처리상태'] == '자동분류']
        df_manual = df_processed_no_duplicates[df_processed_no_duplicates['처리상태'] == '미분류']
        if not df_duplicates.empty:
            with st.expander(f"⚠️ **{len(df_duplicates)}건의 중복 의심 거래가 발견되어 제외됩니다.**"):
                st.dataframe(df_duplicates[['거래일자', '거래내용', '금액']])
        if not df_auto.empty:
            with st.expander(f"🤖 **{len(df_auto)}**건이 자동으로 분류됩니다."):
                df_auto_display = pd.merge(df_auto, data["ACCOUNTS"], on="계정ID", how="left")
                st.dataframe(df_auto_display[['거래일자', '거래내용', '금액', '대분류', '소분류']], hide_index=True)
        col1, col2 = st.columns(2)
        if col1.button("🔙 이전 단계로"):
            st.session_state.current_step = 'upload'
            st.rerun()
        if col2.button("2단계: 분류 작업대 열기 ➡️", type="primary"):
            st.session_state.workbench_data = pd.concat([df_auto, df_manual], ignore_index=True).drop(columns=['duplicate_key'], errors='ignore')
            st.session_state.current_step = 'workbench'
            st.rerun()
    elif st.session_state.current_step == 'workbench':
        if 'workbench_data' not in st.session_state or st.session_state.workbench_data.empty:
            st.success("모든 내역 처리가 완료되었습니다.")
            if st.button("초기 화면으로 돌아가기"):
                st.session_state.current_step = 'upload'
                st.rerun()
            return
        st.subheader(f"✍️ 분류 작업대 (남은 내역: {len(st.session_state.workbench_data)}건)")
        st.info("계정과목이 지정된 항목은 저장 버튼 클릭 시 자동으로 저장됩니다.")
        accounts_df = data["ACCOUNTS"]
        account_options = [""] + [f"[{r['대분류']}/{r['소분류']}] ({r['계정ID']})" for _, r in accounts_df.iterrows()]
        account_map = {f"[{r['대분류']}/{r['소분류']}] ({r['계정ID']})": r['계정ID'] for _, r in accounts_df.iterrows()}
        id_to_account = {v: k for k, v in account_map.items()}
        df_original_workbench = st.session_state.workbench_data.copy()
        df_display = pd.DataFrame()
        df_display['거래일자'] = pd.to_datetime(df_original_workbench['거래일자']).dt.normalize()
        df_display['거래내용'] = df_original_workbench['거래내용']
        df_display['금액'] = df_original_workbench['금액']
        df_display['계정과목_선택'] = df_original_workbench['계정ID'].map(id_to_account).fillna("")
        edited_df = st.data_editor(df_display, hide_index=True, use_container_width=True, key="workbench_editor", disabled=['거래일자', '거래내용', '금액'], num_rows="fixed")
        st.markdown("---")
        with st.expander("✍️ 신규 거래 추가"):
            with st.form("new_transaction_form"):
                c1, c2, c3, c4 = st.columns([2, 4, 2, 3])
                new_date = c1.date_input("거래일자")
                new_desc = c2.text_input("거래내용")
                new_amount = c3.number_input("금액", min_value=0, step=1000)
                new_account = c4.selectbox("계정과목", account_options)
                if st.form_submit_button("추가하기", use_container_width=True):
                    if not all([new_desc, new_amount > 0, new_account]):
                        st.error("거래내용, 금액, 계정과목을 모두 입력해야 합니다.")
                    else:
                        meta = st.session_state.uploaded_file_metadata
                        new_row = {
                            '거래ID': str(uuid.uuid4()), '거래일자': new_date.strftime('%Y-%m-%d'), '사업장명': meta['사업장명'], '구분': meta['구분'],
                            '데이터소스': meta['데이터소스'], '거래내용': new_desc, '금액': new_amount,
                            '계정ID': account_map[new_account], '처리상태': '수동확인'
                        }
                        st.session_state.workbench_data = pd.concat([df_original_workbench, pd.DataFrame([new_row])], ignore_index=True)
                        st.success("새로운 거래가 작업대에 추가되었습니다.")
                        st.rerun()
        st.markdown("---")
        if st.button("💾 저장하기", type="primary"):
            current_state_df = pd.concat([df_original_workbench.drop(columns=['거래일자', '거래내용', '금액', '계정ID']).reset_index(drop=True), edited_df.reset_index(drop=True)], axis=1)
            is_complete = current_state_df['계정과목_선택'].notna() & (current_state_df['계정과목_선택'] != "")
            df_to_process = current_state_df[is_complete].copy()
            df_to_keep = current_state_df[~is_complete].copy()
            if df_to_process.empty:
                st.info("저장할 항목이 없습니다. (계정과목이 지정된 항목이 저장 대상입니다)")
            else:
                df_to_process['계정ID'] = df_to_process['계정과목_선택'].map(account_map)
                original_accounts = df_original_workbench['계정ID'].map(id_to_account).fillna("")
                edited_accounts = df_to_process['계정과목_선택']
                is_changed = original_accounts.reindex(edited_accounts.index) != edited_accounts
                df_to_process.loc[is_changed, '처리상태'] = '수동확인'
                with st.spinner(f"{len(df_to_process)}건의 항목을 저장하는 중입니다..."):
                    final_cols = data["TRANSACTIONS"].columns
                    df_saved = df_to_process.reindex(columns=final_cols).fillna('')
                    combined_trans = pd.concat([data["TRANSACTIONS"], df_saved], ignore_index=True)
                    if update_sheet(SHEET_NAMES["TRANSACTIONS"], combined_trans):
                        st.success(f"{len(df_saved)}건을 성공적으로 저장했습니다.")
                        if df_to_keep.empty:
                            if 'workbench_data' in st.session_state:
                                del st.session_state.workbench_data
                        else:
                            st.session_state.workbench_data = df_original_workbench[df_original_workbench['거래ID'].isin(df_to_keep['거래ID'])].reset_index(drop=True)
                        st.rerun()

def render_settings_page(data):
    st.header("⚙️ 설정 관리")
    tab1, tab2, tab3, tab4 = st.tabs(["🏢 사업장 관리", "📚 계정과목 관리", "🤖 자동분류 규칙", "📄 파일 포맷 관리"])
    with tab1:
        edited_locs = st.data_editor(data["LOCATIONS"], num_rows="dynamic", use_container_width=True, hide_index=True)
        if st.button("사업장 정보 저장", key="save_locations"):
            if update_sheet(SHEET_NAMES["LOCATIONS"], edited_locs): st.success("저장되었습니다."); st.rerun()
    with tab2:
        edited_accs = st.data_editor(data["ACCOUNTS"], num_rows="dynamic", use_container_width=True, hide_index=True)
        if st.button("계정과목 저장", key="save_accounts"):
            if update_sheet(SHEET_NAMES["ACCOUNTS"], edited_accs): st.success("저장되었습니다."); st.rerun()
    with tab3:
        if data["ACCOUNTS"].empty: st.warning("`계정과목 관리` 탭에서 계정과목을 먼저 추가해주세요.")
        else:
            edited_rules = st.data_editor(data["RULES"], num_rows="dynamic", use_container_width=True, hide_index=True,
                column_config={"계정ID": st.column_config.SelectboxColumn("계정ID", options=data["ACCOUNTS"]['계정ID'].tolist(), required=True)})
            if st.button("자동분류 규칙 저장", key="save_rules"):
                if update_sheet(SHEET_NAMES["RULES"], edited_rules): st.success("저장되었습니다."); st.rerun()
    with tab4:
        edited_formats = st.data_editor(data["FORMATS"], num_rows="dynamic", use_container_width=True, hide_index=True,
            column_config={"데이터구분": st.column_config.SelectboxColumn("데이터구분", options=["수익", "비용"], required=True)})
        if st.button("파일 포맷 저장", key="save_formats"):
            if update_sheet(SHEET_NAMES["FORMATS"], edited_formats): st.success("저장되었습니다."); st.rerun()

# =============================================================================
# 5. 메인 실행 로직
# =============================================================================
def main():
    if 'logged_in' not in st.session_state: st.session_state['logged_in'] = False
    if not st.session_state['logged_in']:
        login_screen()
    else:
        st.sidebar.title("🏢 통합 정산 시스템")
        with st.spinner("데이터를 불러오는 중입니다..."):
            data = {name: load_data(sheet) for name, sheet in SHEET_NAMES.items()}
        
        menu = ["📅 월별 정산표", "✍️ 데이터 관리", "⚙️ 설정 관리"]
        choice = st.sidebar.radio("메뉴를 선택하세요.", menu)
        
        st.sidebar.markdown("---")
        if st.sidebar.button("🔃 데이터 새로고침"):
            keys_to_delete = [k for k in st.session_state.keys() if k != 'logged_in']
            for key in keys_to_delete:
                del st.session_state[key]
            st.cache_data.clear()
            st.rerun()

        if st.sidebar.button("로그아웃"): 
            st.session_state.clear()
            st.rerun()
            
        if choice == "📅 월별 정산표": render_pnl_page(data)
        elif choice == "✍️ 데이터 관리": render_data_page(data)
        elif choice == "⚙️ 설정 관리": render_settings_page(data)

if __name__ == "__main__":
    main()
